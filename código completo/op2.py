import tkinter as tk
from tkinter import messagebox
from tkinter.simpledialog import askfloat
import re
from openpyxl import Workbook, load_workbook
import pandas as pd


def alterar_panilha_contrato(name, new_duration, new_cost):
    planilha = load_workbook("planilha.xlsx")
    planilha_main = planilha['members']
    planilha_contrato = planilha['member_ship']
    for linha in planilha_main.iter_rows(values_only=True):
        if name == linha[0]:
            print("Usuário válido")
            age = linha[2]
            Id = linha[1]
            for row in planilha_contrato.iter_rows(values_only=False):
                if Id == row[0].value:
                    row[2].value = new_duration  # Alterar a duração
                    row[3].value = new_cost  # Alterar o custo
                    break

    planilha.save("planilha.xlsx")
    print("contrato altered successfully")


def buscar_membro(name, number):
    planilha = load_workbook("planilha.xlsx")
    planilha_main = planilha['members']
    planilha_contrato = planilha['member_ship']
    valida = False
    member_ship = []

    for linha in planilha_main.iter_rows(values_only=True):
        if name == linha[0] or (number is not None and str(number) == linha[3]):
            print("Usuário válido")
            valida = True
            contrato_valida = False
            account_name = linha[0]
            account_number = linha[3]
            member_ship.extend(linha)
            age = linha[2]
            Id = linha[1]
            for linha_contrato in planilha_contrato.iter_rows(values_only=True):
                if Id == linha_contrato[0]:
                    contrato_valida = True
                    member_ship.append(
                        (linha_contrato[2], linha_contrato[3]))
                    break
            if contrato_valida == False:
                member_ship = None

    if valida:
        return (Id, account_name, age, member_ship)
    else:
        print("Usuário ou número inválido")
        return False


def function2():
    def validar_numeros(entrada):
        if re.match(r'^[0-9]*$', entrada):
            return True
        else:
            return False

    def validar_sem_numeros(entrada):
        if re.search(r'\d', entrada):
            return False
        return True

    def salvar_informacoes():

        nome = entry_nome.get()
        if not nome:
            nome = None
        else:
            nome = nome.strip()

        contato = entry_contato.get()
        if not contato:
            contato = None
        else:
            contato = contato.strip()
            contato = int(contato)

        if nome or contato:
            global resultado
            resultado = buscar_membro(nome, contato)
            if resultado:
                window.destroy()
                segunda_parte(resultado)
                return

            else:
                messagebox.showerror("Falha", "membro não encontrado!")

        else:
            # Aqui você pode adicionar a lógica para salvar as informações em algum lugar
            messagebox.showerror("Falha", "insira os dados nescessários!")

    # Criação da janela
    window = tk.Tk()
    window.title("Cadastro")
    window.geometry("300x150")

    # valida numero e letras
    validar_numeros_cmd = window.register(validar_numeros)
    validar_sem_numeros_cmd = window.register(validar_sem_numeros)

    # Criação dos rótulos
    label_nome = tk.Label(window, text="Nome Completo:")
    label_nome.pack()

    label_contato = tk.Label(window, text="Contato:")
    label_contato.pack()

    # Criação dos campos de entrada
    entry_nome = tk.Entry(window, validate="key",
                          validatecommand=(validar_sem_numeros_cmd, '%P'))
    entry_nome.pack()

    entry_contato = tk.Entry(window, validate="key",
                             validatecommand=(validar_numeros_cmd, '%P'))
    entry_contato.pack()

    # Criação do botão de salvar
    button_salvar = tk.Button(window, text="Buscar",
                              command=salvar_informacoes)

    # butões ocultos

    button_salvar.pack()

    # Iniciar a execução da janela
    window.mainloop()
    return resultado


def segunda_parte(resultado):
    res = resultado

    def exibir_informacoes(res):

        # Aqui você pode substituir os valores abaixo pelos valores reais do usuário
        id_usuario = res[0]
        nome_usuario = res[1]
        idade_usuario = res[2]
        if res[3] != None:
            contrato_complete = res[3]
            contrato_tributação = contrato_complete[4]
            contrato_usuario = f"durabilidade {contrato_tributação[0]}\npreço: {contrato_tributação[1]}"
        else:
            contrato_usuario = "contrato vazio"

        # Exibir as informações na janela
        label_id["text"] = f"ID: {id_usuario}"
        label_nome["text"] = f"Nome: {nome_usuario}"
        label_idade["text"] = f"Idade: {idade_usuario}"
        label_contrato["text"] = contrato_usuario

    # Funções para os botões

    def alterar_contrato():
        new_duration = askfloat("Alterar filiação", "Nova duração:")
        new_cost = askfloat("Alterar filiação", "Novo custo:")
        if new_duration is not None and new_cost is not None:
            messagebox.showinfo("Alterar filiação", "filiação alterado com sucesso!\nNova duração: {}\nNovo custo: {}".format(
                new_duration, new_cost))
            alterar_panilha_contrato(
                res[1], new_duration, new_cost)
            messagebox.showinfo("Contrato alterado")

            window.destroy()
        else:
            messagebox.showwarning("Alterar filiação", "Valores inválidos.")

    def deletar_contrato():
        delete_contrato(res[1])
        window.destroy()

    def delete_contrato(member_name):
        planilha = load_workbook("planilha.xlsx")
        planilha_membership = planilha['member_ship']
        linha_encontrada_membership = None

        # Excluir na planilha 'member_ship'
        for i, linha_membership in enumerate(planilha_membership.iter_rows(min_row=2, values_only=True), start=2):
            if member_name == linha_membership[1]:
                linha_encontrada_membership = i
                print("Encontrado na linha em 'member_ship':",
                      linha_encontrada_membership)
                planilha_membership.delete_rows(linha_encontrada_membership)
                break  # Sai do loop após encontrar a primeira correspondência

        planilha.save("planilha.xlsx")  # Salva as alterações na planilha
        messagebox.showinfo("Deletar filiação",
                            "filiação deletado com sucesso!")

    def deletar_membro():
        delete_member(res[1])

        window.destroy()

    # Criação da janela
    window = tk.Tk()
    window.title("Informações do Usuário")

    # Criação dos rótulos para as informações do usuário
    label_id = tk.Label(window)
    label_id.pack()

    label_nome = tk.Label(window)
    label_nome.pack()

    label_idade = tk.Label(window)
    label_idade.pack()

    label_contrato = tk.Label(window)
    label_contrato.pack()

    # Definir tamanho dos botões
    button_width = 20
    button_height = 2

    # Criação dos botões
    button_alterar_contrato = tk.Button(
        window, text="Alterar contrato", width=button_width, height=button_height, command=alterar_contrato)
    button_alterar_contrato.pack()

    button_deletar_contrato = tk.Button(
        window, text="Deletar contrato", width=button_width, height=button_height, command=deletar_contrato)
    button_deletar_contrato.pack()

    button_deletar_membro = tk.Button(
        window, text="Deletar Membro", width=button_width, height=button_height, command=deletar_membro)
    button_deletar_membro.pack()

    # Exibir as informações do usuário na janela
    exibir_informacoes(res)

    # Iniciar a execução da janela
    window.mainloop()


def delete_member(member_name):
    planilha = load_workbook("planilha.xlsx")
    planilha_member = planilha['members']
    planilha_membership = planilha['member_ship']
    linha_encontrada_member = None
    linha_encontrada_membership = None

    # Excluir na planilha 'members'
    for i, linha_member in enumerate(planilha_member.iter_rows(min_row=2, values_only=True), start=2):
        if member_name == linha_member[0]:
            linha_encontrada_member = i
            print("Encontrado na linha em 'members':", linha_encontrada_member)
            planilha_member.delete_rows(linha_encontrada_member)
            break  # Sai do loop após encontrar a primeira correspondência

    # Excluir na planilha 'member_ship'
    for i, linha_membership in enumerate(planilha_membership.iter_rows(min_row=2, values_only=True), start=2):
        if member_name == linha_membership[1]:
            linha_encontrada_membership = i
            print("Encontrado na linha em 'member_ship':",
                  linha_encontrada_membership)
            planilha_membership.delete_rows(linha_encontrada_membership)
            break  # Sai do loop após encontrar a primeira correspondência

    planilha.save("planilha.xlsx")  # Salva as alterações na planilha
