import string
import random
from openpyxl import Workbook, load_workbook
import pandas as pd
import os

# Verificar e instalar bibliotecas ausentes


def criar_planilha():
    if os.path.exists("planilha.xlsx"):
        print("Planilha already exists")
        return

    dados = {'name': [],
             'id': [],
             'age': [],
             'contacts': []}
    dados_history = {"id": [],
                     'name': [],
                     'duration': [],
                     'cost': []}
    df_main = pd.DataFrame(dados)

    df_history = pd.DataFrame(dados_history)
    with pd.ExcelWriter("planilha.xlsx") as writer:
        df_main.to_excel(writer, sheet_name='members', index=False)
        df_history.to_excel(writer, sheet_name='member_ship', index=False)

    print("Planilhas created successfully.")


def create_account(name, age, number, duration, cost):
    df_main = pd.read_excel("planilha.xlsx", sheet_name='members')
    df_history = pd.read_excel("planilha.xlsx", sheet_name='member_ship')
    balance = 1
    account = gerar_sequencia()
    nova_conta_main = pd.DataFrame({'name': [name],
                                    'id': [account],
                                    'age': [age],
                                    'contacts': [number]})

    nova_conta_history = pd.DataFrame({'id': [account],
                                       'name': [name],
                                       'duration': [duration],
                                      'cost': [cost]})

    # Adiciona a nova conta ao DataFrame 'main'
    df_main = pd.concat(
        [df_main, nova_conta_main], ignore_index=False)
    df_history = pd.concat(
        [df_history, nova_conta_history], ignore_index=False)

    # Salva os DataFrames atualizados no arquivo Excel
    with pd.ExcelWriter("planilha.xlsx") as writer:
        df_main.to_excel(
            writer, sheet_name='members', index=False)
        df_history.to_excel(
            writer, sheet_name='membership', index=False)

    print("Account created successfully")


def gerar_sequencia():
    # Gera uma sequência aleatória de duas letras
    letras = random.choices(string.ascii_uppercase, k=2)

    # Gera um número de 6 dígitos
    numero_6_digitos = random.randint(100000, 999999)

    # Gera um número aleatório de 1 a 9
    numero_1_a_9 = random.randint(1, 9)

    # Retorna a sequência gerada
    return ''.join(letras) + str(numero_6_digitos) + "RASTR0" + str(numero_1_a_9)


def validate_login(name, password):
    planilha = load_workbook("planilha.xlsx")
    planilha_ativa = planilha['main']
    user = False
    valida = False
    account_name = None
    account_number = None
    balance = None

    for linha in planilha_ativa.iter_rows(min_row=2, values_only=True):
        if name == linha[0]:
            user = True
            print(f"Valid user")
            if password == linha[2]:
                print(f"Valid password")
                valida = True
                account_name = linha[0]
                account_number = linha[3]
                balance = linha[4]
                break

    if not user:
        print("Invalid username")

    elif not valida:
        print("Invalid password")

    return valida, account_name, account_number, balance

    planilha = load_workbook("planilha.xlsx")
    main = planilha["main"]
    number = f"{account_number}"
    encontrado = False

    for celula in main["D"]:
        if number == celula.value:
            encontrado = True
            linha = celula.row
            balance = main[f'E{linha}'].value
            balance += deposit
            main[f"E{linha}"] = balance
            print(f"New Balance: {balance}")
            print("Deposit successfully")
            reg = f"Deposit {deposit}$ New balance: {balance}"
            planilha.save("planilha.xlsx")

            break

    if not encontrado:
        print("Não encontrado")
    planilha = load_workbook("planilha.xlsx")
    aba_history = planilha['history']
    coluna_a = aba_history['A']

    for celula in coluna_a:
        if celula.value == account_number:
            linha = celula.row
            new_history = aba_history[f'B{linha}'].value + "/" + reg
            aba_history[f'B{linha}'].value = new_history
            planilha.save("planilha.xlsx")

    planilha = load_workbook("planilha.xlsx")
    history = planilha['history']
    for number in history['A']:
        if account_number == number.value:
            linha = number.row
            print(yellow, history[f"B{linha}"].value)

    planilha = load_workbook("planilha.xlsx")
    main = planilha["main"]
    number = f"{account_number}"
    encontrado = False

    for celula in main["D"]:
        if number == celula.value:
            encontrado = True
            linha = celula.row
            balance = main[f'E{linha}'].value
            if withdraw > balance:
                print("You don't have enough!")
            balance -= withdraw
            main[f"E{linha}"] = balance
            print(f"New Balance: {balance}")
            print("withdrawal successfully")

            # Atualizando historico

            reg = f" {withdraw}$ withdrawn, New balance: {balance}"
            break

    if not encontrado:
        print(f"Não encontrado")
    planilha.save("planilha.xlsx")
    aba_history = planilha['history']
    coluna_a = aba_history['A']

    for celula in coluna_a:
        if celula.value == account_number:
            linha = celula.row
            new_history = aba_history[f'B{linha}'].value + "/" + reg
            aba_history[f'B{linha}'].value = new_history
            planilha.save("planilha.xlsx")


def delete_account(account_number):
    planilha = load_workbook("planilha.xlsx")
    main = planilha['main']
    find = False
    for linha in main['D']:
        if account_number == linha.value:
            linha_da_conta = linha.row
            main.delete_rows(linha_da_conta)
            planilha.save("planilha.xlsx")
            print("Account deleted!")
            find = True
    if find == False:
        print("Account not found!")
    reg = "Account deleted"
    aba_history = planilha['history']
    coluna_a = aba_history['A']

    for celula in coluna_a:
        if celula.value == account_number:
            linha = celula.row
            new_history = aba_history[f'B{linha}'].value + "/" + reg
            aba_history[f'B{linha}'].value = new_history
            planilha.save("planilha.xlsx")

    print("=" * 20)
    valid = ["1", "2", "3", "4", "5", "6", "", " "]
    m = input(
        f"[Enter] To Close\n[1] Create Account\n[2] Login\n[3] Deposit\n[4] Withdrawal\n[5] Help\n\n: ")
    while m not in valid:
        print("Invalid option!")
        m = input("[1] Create Account\n[2] Login\n[3] Close\n: ")

    return m

# Função principal

    criar_planilha()
    while True:
        m = menu()
        if m == '1':
            create_account()
        elif m == '2':
            print("Login to your account")
            name = str(input("Name: "))
            password = str(input("Password: "))

            valida, account_name, account_number, balance = validate_login(
                name, password)

            if valida == True:
                account_number = f"{account_number}"
                print(f"Login successful!")
                print("Account Information:")
                print("Name:", account_name)
                print("Account:", account_number)
                print("Balance:", balance)

                while True:
                    print("=" * 20)
                    menu_option = input(
                        "[1] Deposit\n[2] Withdrawal\n[3] History\n[4] Delete\n[5] Close\n: ")

                    if menu_option == '1':
                        deposit = float(input("Deposit amount: "))
                        deposit_amount(account_number, deposit)
                    elif menu_option == '2':
                        withdraw = float(input("Withdraw amount: "))
                        withdrawal_amount(account_number, withdraw)
                    elif menu_option == '3':
                        show_history(account_number)
                    elif menu_option == '4':
                        delete_account(account_number)
                        break
                    elif menu_option == '5':
                        break
                    else:
                        print("Invalid option!")
            else:
                print(f"Login failed!.")

        elif m == '3':
            account_number = input("Account number: ")
            planilha = load_workbook("planilha.xlsx")
            planilha_ativa = planilha['main']
            user = False
            for linha in planilha_ativa["D"]:
                if account_number == linha.value:
                    user = True
            if user == True:
                cash = float(input("Value: "))
                deposit_amount(account_number, cash)

            else:
                print("Account not found")

        elif m == '4':
            account_number = input("Account number: ")
            planilha = load_workbook("planilha.xlsx")
            planilha_ativa = planilha['main']
            user = False
            for linha in planilha_ativa["D"]:
                if account_number == linha.value:
                    user = True
                    value = planilha_ativa.cell(row=linha.row, column=5).value

            if user == True:
                cash = float(input("Value: "))
                if cash <= value:
                    withdrawal_amount(account_number, cash)
                else:
                    print("Insufficient funds")

        elif m == '5':
            print(f"""You can get help by contacting the link below:
https://api.whatsapp.com/send?hone=996528343

or via email: dscontac@hotmail.com

code: github.com/diegimon""")
        # elif m == '6':
        elif m == '' or m == ' ':
            print(
                f"Program finished thanks for the preference!")
            finaly = input("[press any button to end]\n")
            break
