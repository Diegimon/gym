import tkinter as tk
from tkinter import messagebox
from tkinter.simpledialog import askfloat
import re
from openpyxl import Workbook, load_workbook
import pandas as pd


def alterar_panilha_membership(name, new_duration, new_cost):
    planilha = load_workbook("planilha.xlsx")
    planilha_main = planilha['members']
    planilha_membership = planilha['membership']
    valida = False

    for linha in planilha_main.iter_rows(values_only=True):
        if name == linha[0]:
            print("Usuário válido")
            valida = True
            account_name = linha[0]
            account_number = linha[3]
            age = linha[2]
            Id = linha[1]
            for row in planilha_membership.iter_rows(values_only=False):
                if Id == row[0].value:
                    row[2].value = new_duration  # Alterar a duração
                    row[3].value = new_cost  # Alterar o custo
                    break

    planilha.save("planilha.xlsx")
    print("Membership altered successfully")


def show_buttons():
    button1.config(state="normal")
    button2.config(state="normal")
    button3.config(state="normal")


def buscar_membro(name, number):
    number = str(number)
    planilha = load_workbook("planilha.xlsx")
    planilha_main = planilha['members']
    planilha_membership = planilha['membership']
    valida = False
    member_ship = []

    for linha in planilha_main.iter_rows(values_only=True):
        if name == linha[0] or number == linha[3]:
            print("Usuário válido")
            valida = True
            account_name = linha[0]
            account_number = linha[3]
            member_ship.extend(linha)
            age = linha[2]
            Id = linha[1]
            for linha_membership in planilha_membership.iter_rows(values_only=True):
                if Id == linha_membership[0]:
                    member_ship.append(
                        (linha_membership[2], linha_membership[3]))
                    break

    if valida:
        return (Id, account_name, age, member_ship)
    else:
        print("Usuário ou número inválido")
        return False


def function2():
    def validar_numeros(entrada):
        if re.match(r'^[0-9]*$', entrada):
            return True
        else:
            return False

    def validar_sem_numeros(entrada):
        if re.search(r'\d', entrada):
            return False
        return True

    def salvar_informacoes():

        nome = entry_nome.get()
        if not nome:
            nome = None
        else:
            nome = nome.strip()

        contato = entry_contato.get()
        if not contato:
            contato = None
        else:
            contato = contato.strip()
            contato = int(contato)

        if nome or contato:
            global resultado
            resultado = buscar_membro(nome, contato)
            if resultado:
                window.destroy()
                segunda_parte(resultado)
                return

            else:
                messagebox.showerror("Falha", "membro não encontrado!")

        else:
            # Aqui você pode adicionar a lógica para salvar as informações em algum lugar
            messagebox.showerror("Falha", "insira os dados nescessários!")

    # Criação da janela
    window = tk.Tk()
    window.title("Cadastro")
    window.geometry("300x150")

    # valida numero e letras
    validar_numeros_cmd = window.register(validar_numeros)
    validar_sem_numeros_cmd = window.register(validar_sem_numeros)

    # Criação dos rótulos
    label_nome = tk.Label(window, text="Nome Completo:")
    label_nome.pack()

    label_contato = tk.Label(window, text="Contato:")
    label_contato.pack()

    # Criação dos campos de entrada
    entry_nome = tk.Entry(window, validate="key",
                          validatecommand=(validar_sem_numeros_cmd, '%P'))
    entry_nome.pack()

    entry_contato = tk.Entry(window, validate="key",
                             validatecommand=(validar_numeros_cmd, '%P'))
    entry_contato.pack()

    # Criação do botão de salvar
    button_salvar = tk.Button(window, text="Buscar",
                              command=salvar_informacoes)

    # butões ocultos

    button_salvar.pack()

    # Iniciar a execução da janela
    window.mainloop()
    return resultado


def segunda_parte(resultado):
    res = resultado

    def exibir_informacoes(res):

        # Aqui você pode substituir os valores abaixo pelos valores reais do usuário
        id_usuario = res[0]
        nome_usuario = res[1]
        idade_usuario = res[2]
        membership_complete = res[3]
        membership_tributação = membership_complete[4]
        membership_usuario = f"durabilidade {membership_tributação[0]}\npreço: {membership_tributação[1]}"

        # Exibir as informações na janela
        label_id["text"] = f"ID: {id_usuario}"
        label_nome["text"] = f"Nome: {nome_usuario}"
        label_idade["text"] = f"Idade: {idade_usuario}"
        label_membership["text"] = membership_usuario

    # Funções para os botões

    def alterar_membership():
        new_duration = askfloat("Alterar filiação", "Nova duração:")
        new_cost = askfloat("Alterar filiação", "Novo custo:")
        if new_duration is not None and new_cost is not None:
            messagebox.showinfo("Alterar filiação", "filiação alterado com sucesso!\nNova duração: {}\nNovo custo: {}".format(
                new_duration, new_cost))
            alterar_panilha_membership(
                res[1], new_duration, new_cost)
        else:
            messagebox.showwarning("Alterar filiação", "Valores inválidos.")

    def deletar_membership():
        messagebox.showinfo("Deletar filiação",
                            "filiação deletado com sucesso!")

    def deletar_membro():
        messagebox.showinfo("Deletar Membro", "Membro deletado com sucesso!")

    # Criação da janela
    window = tk.Tk()
    window.title("Informações do Usuário")

    # Criação dos rótulos para as informações do usuário
    label_id = tk.Label(window)
    label_id.pack()

    label_nome = tk.Label(window)
    label_nome.pack()

    label_idade = tk.Label(window)
    label_idade.pack()

    label_membership = tk.Label(window)
    label_membership.pack()

    # Definir tamanho dos botões
    button_width = 20
    button_height = 2

    # Criação dos botões
    button_alterar_membership = tk.Button(
        window, text="Alterar Membership", width=button_width, height=button_height, command=alterar_membership)
    button_alterar_membership.pack()

    button_deletar_membership = tk.Button(
        window, text="Deletar Membership", width=button_width, height=button_height, command=deletar_membership)
    button_deletar_membership.pack()

    button_deletar_membro = tk.Button(
        window, text="Deletar Membro", width=button_width, height=button_height, command=deletar_membro)
    button_deletar_membro.pack()

    # Exibir as informações do usuário na janela
    exibir_informacoes(res)

    # Iniciar a execução da janela
    window.mainloop()


def delete_member(account_number):
    planilha = load_workbook("planilha.xlsx")
    planilha_member = planilha['members']
    planilha_membership = planilha['membership']
    for linha_member in planilha_member:
        if account_number == planilha_member[1]:
            print("Encontrado!")
