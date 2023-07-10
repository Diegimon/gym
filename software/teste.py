from openpyxl import load_workbook


def delete_member(account_number):
    planilha = load_workbook("planilha.xlsx")
    planilha_member = planilha['members']
    linha = 0

    for row in planilha_member.iter_rows(min_row=0, values_only=True):
        if row[1] == account_number:
            print(row[1])
            planilha_member.delete_rows(planilha_member, linha)
            break

    else:
        linha += 1
        print("Número da conta não encontrado!")

    planilha.save("planilha.xlsx")


delete_member("MT960055RASTR03")
