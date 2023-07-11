from tkinter import *
from openpyxl import load_workbook


def mostrar_itens_tkinter():
    planilha = load_workbook("planilha.xlsx")
    planilha_member = planilha['members']
    planilha_member_ship = planilha['member_ship']

    # Criar janela
    window = Tk()
    window.title("Valores da Planilha")

    # Criar Text Widget para exibir os valores
    text_widget = Text(window, height=10, width=50)
    text_widget.pack()

    # Obter todos os valores da planilha 'members'
    valores_members = [linha for linha in planilha_member.iter_rows(
        min_row=2, values_only=True)]
    for linha_member in valores_members:
        text_widget.insert(END, "membro:\n")
        for valor in linha_member:
            text_widget.insert(END, f"{valor} ")
        text_widget.insert(END, "\n\n")

        # Procurar linhas correspondentes na planilha 'member_ship'
        for linha_ship in planilha_member_ship.iter_rows(min_row=2, values_only=True):
            if linha_member[0] == linha_ship[1]:
                text_widget.insert(END, "Contrato:\n")
                for valor in linha_ship:
                    text_widget.insert(END, f"{valor} ")
                text_widget.insert(END, "\n\n")

    # Executar o loop principal do Tkinter
    window.mainloop()
